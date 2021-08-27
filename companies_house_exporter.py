# companies_house_exporter.py

""" UK Companies House viewer - use the UK companies house to get details on multiple companeis into an Excel file

The script allows the user to provide an Excel with UK company number and add the current registration details
of these companies into the same Excel file through the Companies House API

In order to use this tool, the user must obtain an API key from https://developer-specs.company-information.service.gov.uk/guides/authorisation
the API key should be saved in a text file named api_key.txt.
In addition, an Excel file containing the numbers of the relevant companies in the first column must be provided.

The script imports the ch_api.py module, that include calls to the Companies House API.

Created: Shai Shulman, 2021, under the GNU General Public License v3
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from ch_api import get_profile, get_officers
from datetime import datetime, timedelta
import argparse

XL_FILE = 'Companies.xlsx'
API_KEY_FILE = 'api_key.txt'
DATE_FIELDS = ['next_confirmation', 'next_accounts']
ALERT_DAYS_FOR_DATES = 10


def _parse_args():
    parser = argparse.ArgumentParser(description='Get registration details for companies from the UK Companies House '
                                                 'and save in Excel file')
    parser.add_argument('-f', '--file', type=str, default=XL_FILE, help='Excel file with registration numbers')
    parser.add_argument('-k', '--key', type=str, default=None, help='Key for Companies House API (if empty - will use '
                                                                    'api_key.txt file')
    parser.add_argument('-n', '--nextdate', action='store_true', required=False,
                        help='Get next date for validation statement')
    parser.add_argument('-d', '--directors', action='store_true', required=False, help='Get names of directors')
    return parser.parse_args()


def get_sheet(filename):
    """
    Open workbook and return first worksheet as object

    Args:
        filename: local file name

    Returns:
        Worksheet object

    """
    wb = load_workbook(filename)
    return wb.worksheets[0]


def get_first_data_row(sheet):
    """ Find first row in the excel with data (company numbers)

    Args:
        sheet: worksheet object

    Returns:
        first row with data

    """
    for row in range(1, sheet.max_row):
        if str(sheet['A' + str(row)].value).isnumeric():
            return row
    return 1


def get_numbers_from_xl(sheet, data_row):
    """Get company numbers from the first column of the excel sheet

    Args:
        sheet: worksheet object
        data_row: first row containing data

    Returns:
        list of company numbers

    """
    numbers = []
    for row in range(data_row, sheet.max_row + 1):
        numbers.append(sheet['A' + str(row)].value)

    return numbers


def get_api_key_from_file():
    """ Open the text file named API_KEY_FILE  and read the api key

    Returns:
        API key as string

    """
    try:
        f = open(API_KEY_FILE, 'r')
        akey = f.readline()
        f.close()
        return akey
    except FileNotFoundError:
        print(f'ERROR: file {API_KEY_FILE} not found')
    except:
        print(f'ERROR: cannot read api key from file {API_KEY_FILE}')


def get_companies_data(numbers, key, include_directors=False):
    """Get data of companies from the Companies House

    Args:
        numbers: list of numbers of companies to search
        include_directors: should the query include the names of current directors

    Returns:

    """
    results = []
    for num in numbers:
        print(f'Obtaining data for company number {num}...', end='')
        profile = get_profile(num, key)
        if profile:
            if include_directors:
                profile['directors'] = get_officers(num, key, True)
            profile['company_num'] = num
            print('done')
        else:
            profile = {'company_num': num}
            print('failed')
        results.append(profile)
    return results


def save_xl_sheet(sheet, data, first_data_row, file_name):
    """Save companies data into the Excel worksheet, in the same row where existing company numbers are located

    Args:
        sheet: worksheet object
        data: list of dictionary objects that include the data to be saved, by same order of rows as included in the Excel
        first_data_row: first row containing data
        file_name: name of file to be saved
    """
    data_fields = {'company_name': 2, 'address': 3, 'next_confirmation': 4, 'next_accounts': 5, 'directors': 6}
    print(f'Saving excel to {file_name}...', end='')
    for row in range(first_data_row, len(data) + first_data_row):
        for field_name, col in data_fields.items():
            sheet.cell(row, col).value = data[row - first_data_row].get(field_name, '')
            if field_name in DATE_FIELDS and data[row - first_data_row].get(field_name, None):
                if data[row - first_data_row].get(field_name, None) < datetime.today().date() + timedelta(days=90):
                    sheet.cell(row, col).fill = PatternFill("solid", "ff0000")
    sheet.parent.save(file_name)
    print('done')


if __name__ == "__main__":
    args = _parse_args()
    key = args.key if args.key else get_api_key_from_file()
    xl_sheet = get_sheet(args.file)
    first_row = get_first_data_row(xl_sheet)
    data = get_companies_data(get_numbers_from_xl(xl_sheet, first_row), key, args.directors)
    save_xl_sheet(xl_sheet, data, first_row, args.file)
