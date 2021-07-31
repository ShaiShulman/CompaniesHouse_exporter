# ch_filing_explorer.py

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from api import get_profile, get_officers
from datetime import datetime, timedelta
import argparse

XL_FILE = 'Companies.xlsx'
API_KEY_FILE = 'api_key.txt'
DATE_FIELDS = ['next_confirmation', 'next_accounts']
ALERT_DAYS_FOR_DATES = 10

def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', type=str, default=XL_FILE, help='Excel file with registration numbers')
    parser.add_argument('-k', '--key', type=str, default=None, help='Key for Companies House API (if empty - will use '
                                                                    'api_key.txt file')
    parser.add_argument('-n', '--nextdate', action='store_true', required=False,
                        help='Get next date for validation statement')
    parser.add_argument('-d', '--directors', action='store_true', required=False, help='Get names of directors')
    parser.add_argument('-o', '--officers', action='store_true', required=False, help='Get names of directors')
    return parser.parse_args()


def get_sheet(filename):
    ''' load excel worksheet '''
    wb = load_workbook(filename)
    return wb.worksheets[0]


def get_first_data_row(sheet):
    for row in range(1, sheet.max_row):
        if str(sheet['A' + str(row)].value).isnumeric():
            return row
    return 1


def get_numbers_from_xl(sheet, data_row):
    numbers = []
    for row in range(data_row, sheet.max_row+1):
        numbers.append(sheet['A' + str(row)].value)

    return numbers


def get_api_key_from_file():
    try:
        f = open(API_KEY_FILE, 'r')
        akey = f.readline()
        f.close()
        return akey
    except FileNotFoundError:
        print(f'ERROR: file {API_KEY_FILE} not found')
    except:
        print(f'ERROR: cannot read api key from file {API_KEY_FILE}')


def get_companies_data(numbers, include_directors=False, include_address=False):
    results = []
    for num in numbers:
        print(f'Obtaining data for company number {num}...', end='')
        profile = get_profile(num)
        if profile:
            if include_directors:
                profile['directors'] = get_officers(num, True)
            profile['company_num'] = num
            print('done')
        else:
            profile = {'company_num': num}
            print('failed')
        results.append(profile)
    return results


def save_xl_sheet(sheet, data, first_data_row, file_name):
    data_fields = {'company_name': 2, 'address': 3, 'next_confirmation': 4, 'next_accounts': 5, 'directors': 6}
    print(f'Saving excel to {file_name}...', end='')
    for row in range(first_data_row, len(data) + first_data_row):
        for field_name, col in data_fields.items():
            sheet.cell(row, col).value = data[row-first_data_row].get(field_name, '')
            if field_name in DATE_FIELDS and data[row-first_data_row].get(field_name, None):
                if data[row-first_data_row].get(field_name, None) < datetime.today().date() + timedelta(days=90):
                    sheet.cell(row, col).fill = PatternFill("solid", "ff0000")
    sheet.parent.save(file_name)
    print('done')


if __name__ == "__main__":
    args = _parse_args()
    key = args.key if args.key else get_api_key_from_file()
    xl_sheet = get_sheet(args.file)
    first_row = get_first_data_row(xl_sheet)
    data = get_companies_data(get_numbers_from_xl(xl_sheet, first_row), args.directors)
    save_xl_sheet(xl_sheet, data, first_row, args.file)



