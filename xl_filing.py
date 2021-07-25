# xl_filing.py 

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from api import get_next_confirmation_date
from datetime import datetime, timedelta
import argparse

XL_FILE = "Entities.xlsx"


def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', type=str, default=XL_FILE, help='Excel file with registration numbers')
    parser.add_argument('-k', '--key', type=str, default=None, help='Key for Companies House API (if empty - will use '
                                                                    'api.txt file')
    parser.add_argument('-n', '--nextdate', action='store_true', required=False,
                        help='Get next date for validation statement')
    parser.add_argument('-d', '--directors', action='store_true', required=False, help='Get names of directors')
    parser.add_argument('-o', '--officers', action='store_true', required=False, help='Get names of directors')
    return parser.parse_args()


def get_sheet():
    wb = load_workbook(filename=XL_FILE)
    return wb['Sheet']


def get_numbers_from_xl():
    numbers = []
    ws = get_sheet()
    for row in range(2, ws.max_row):
        numbers.append(ws['A' + str(row)].value)

    return get_next_confirmation_date(numbers)


def update_sheet_nextdate():
    print("Getting list of entities")
    reg_numbers = get_numbers_from_xl()
    print("Getting confirmation dates")
    entities = get_next_confirmation_date(reg_numbers)
    print("updating excel file")
    ws = get_sheet()

    for number, date in entities.items():
        for row in range(2, ws.max_row):
            if ws['A' + str(row)].value == number:
                ws['D' + str(row)].value = date
                if datetime.strptime(date, "%Y-%m-%d").date() < datetime.today().date() + timedelta(days=90):
                    ws['D' + str(row)].fill = PatternFill("solid", "ff0000")
    ws.parent.save(XL_FILE)
    print(f"{XL_FILE_DEFAULT} saved")


if __name__ == "__main__":
    args = _parse_args()

    if args.nextdate:
        update_sheet_nextdate()
